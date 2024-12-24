# Imports
from openpyxl import load_workbook
from openpyxl import Workbook

# Open the workbook
wb = load_workbook(filename='excelFiles/sells.xlsx')

# Access the specific sheet by name
sheet = wb['sells']

# Print the value of cell A2
for row in range(2, 7):  # A2 to A6 corresponds to rows 2 to 6
    cell_value = sheet[f'A{row}'].value  # Get value of cell A{row}
    print(f'Value in A{row}: {cell_value}')  # Print value of the cell

# We can also add the value of the cells to an array
products = []
for row in range(2, 7):
    products.append(sheet[f'A{row}'].value)

print(products);

# Finally we can make another file with the list

# Create the new file
book = Workbook()
newSheet = book.active

# Iterate over the products list and write each element to the new sheet
for index, element in enumerate(products, start=1):
    newSheet[f'A{index}'] = element

# Save the new workbook
book.save("products.xlsx")
