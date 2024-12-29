# We are going to use CompareTwoFiles.py as a base and then 
# we are going to add the functionality to write down the coincidences in a new file

# Imports
from openpyxl import load_workbook
from openpyxl import Workbook 
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Open the first workbook (the one with all the list)
wb = load_workbook(filename='excelFiles/drinks.xlsx')

# Access the specific sheet by name
sheet = wb['Hoja1']

# get the list of products to find in the column
drinksList = []

# Iterate and append until findin a empty value
for cell in sheet['A'][1:]:
    drinksList.append(cell.value);
    if cell.value is None:
        break

# Open the file with all the data
wb = load_workbook(filename='excelFiles/sells.xlsx')
sheet = wb['sells']

# Create the data array
data = []
coincidences = 0

# make the loop to find the coincidences and add the information into the array
for index,cell in enumerate(sheet["A"][1:], start=2): # make sure to start the index at 2 so we can ignore the title of the table
    if cell.value in drinksList:

        # Make sure to add a list for every coincidence
        data.append([])

        # Append all the data
        data[coincidences].append(sheet[f"A"+str(index)].value)
        data[coincidences].append(sheet[f"B"+str(index)].value)
        data[coincidences].append(sheet[f"C"+str(index)].value)

        coincidences += 1

    if cell.value is None:
        break # Break when finding a empty cell

# Create the new file
book = Workbook()
sheet = book.active

# Make the format for the new file
sheet["A1"] = "Producto"
sheet["B1"] = "Cantidad"
sheet["C1"] = "Precio unitario"

# Put design to the format
sheet["A1"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
sheet["B1"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
sheet["C1"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
 
# Now we iterate on the list to writed down
for index,element in enumerate(data, start=2):
    sheet[f"A"+str(index)] = element[0]
    sheet[f"B"+str(index)] = element[1]
    sheet[f"C"+str(index)] = element[2]
    print(element)

# Save the file
book.save("coincidences.xlsx")