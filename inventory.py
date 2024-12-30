# We want to know how much of a product we need to buy base on a inventory file 
# and a ideal inventory file, finally the code is 
# going to make file showing the info
# also we are going to consider how much items are in a box


# We are going to use the CTFAndWrite.py file as a base

# Imports
from openpyxl import load_workbook
from openpyxl import Workbook 
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

# Open the first workbook (the one with the ideals)
wb = load_workbook(filename='excelFiles/ideal.xlsx')
sheet = wb['Hoja1']

# make the list of products to find and the all data 2D array
productList = []
allData = []

# Fill the list

# Iterate and append until findin a empty value
for cell in sheet['A'][1:]:
    if cell.value is None:
        break
    productList.append(cell.value);

# Do the same but with all the Data
for index,cell in enumerate(sheet['A'][1:], start=2):
    if cell.value is None:
        break

    allData.append([])
    
