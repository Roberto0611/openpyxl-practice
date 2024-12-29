# In this file we are going to compare two excel column and then
# based on the coincidence get the information of another column and add it up
# We are going to get the summation of the quantity column of the coincidence

# Imports
from openpyxl import load_workbook
from openpyxl import Workbook 

# Open the first workbook (the one with all the list)
wb = load_workbook(filename='excelFiles/drinks.xlsx')

# Access the specific sheet by name
sheet = wb['Hoja1']

# get the list of products to find in the column
drinksList = []

# Iterate and append until findin a empty value
for cell in sheet['A'][1:]:
    drinksList.append(cell.value);
    if cell.value is None:
        break

# Open the file with all the data
wb = load_workbook(filename='excelFiles/sells.xlsx')
sheet = wb['sells']

# Create the quantity var
quantity = 0

# make the loop to find the coincidences and add it up
for index,cell in enumerate(sheet["A"][1:], start=2): # make sure to start the index at 2 so we can ignore the title of the table
    if cell.value in drinksList:
        # get the quantity of the coincidence and ad it up
        quantity += sheet[f"B"+str(index)].value

    if cell.value is None:
        break # Break when finding a empty cell

# Finally show the final quantity
print(quantity)